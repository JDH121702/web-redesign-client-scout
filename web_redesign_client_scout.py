import ast

import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import numpy as np
import time
import random
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import sys
from contextlib import contextmanager
from typing import Dict, Iterable, List, Sequence
from urllib.parse import urlparse
import re

import requests
from bs4 import BeautifulSoup

# Set page config
st.set_page_config(page_title="Web Redesign Client Scout", layout="wide", initial_sidebar_state="expanded")

# Utility helpers -----------------------------------------------------------

def get_asset_path(relative_path: str) -> Path:
    """Return an absolute path for bundled assets.

    When the application is bundled as a PyInstaller executable, assets are
    copied to a temporary directory exposed through ``sys._MEIPASS``. During
    normal development ``__file__`` is available instead. This helper keeps the
    rest of the code agnostic of the execution environment.
    """

    base_path = getattr(sys, "_MEIPASS", Path(__file__).parent)  # type: ignore[attr-defined]
    return Path(base_path) / relative_path


@contextmanager
def styled_card(class_name: str = "dashboard-card"):
    """Context manager that wraps content in a styled card container."""

    st.markdown(f"<div class='{class_name}'>", unsafe_allow_html=True)
    try:
        yield
    finally:
        st.markdown("</div>", unsafe_allow_html=True)


# Data model and helper utilities -------------------------------------------

CLIENT_DATA_SCHEMA: Dict[str, str] = {
    "Company Name": "object",
    "Website URL": "object",
    "Industry": "object",
    "Contact Person": "object",
    "Contact Email": "object",
    "Contact Phone": "object",
    "Last Website Update": "datetime64[ns]",
    "Mobile Friendly": "object",
    "Website Speed Score": "float64",
    "Design Score": "float64",
    "Potential Value": "float64",
    "Priority": "object",
    "Notes": "object",
    "Last Contact Date": "datetime64[ns]",
    "Status": "object",
    "Design Summary": "object",
    "Design Strengths": "object",
    "Design Gaps": "object",
    "Design Recommendations": "object",
    "Design Breakdown": "object",
}

DESIGN_CATEGORY_LIBRARY: Dict[str, Dict[str, str]] = {
    "Brand Cohesion": {
        "strength": "Visual identity feels consistent and premium across the journey.",
        "gap": "Brand cues shift between sections, diluting trust and recognition.",
        "action": "Audit typography, colors, and imagery to create a documented UI kit that locks the brand together.",
    },
    "Visual Hierarchy": {
        "strength": "Key sections guide the eye smoothly with clear spacing and typography.",
        "gap": "Important messages compete for attention, forcing visitors to hunt for next steps.",
        "action": "Restructure hero and pillar sections so that one primary action is obvious on every screen.",
    },
    "Content Clarity": {
        "strength": "Messaging is concise and scannable, making the value proposition easy to grasp.",
        "gap": "Copy blocks are dense, and supporting facts are buried below the fold.",
        "action": "Rewrite key pages with skimmable headings, proof points, and simplified copy.",
    },
    "Conversion Readiness": {
        "strength": "Calls-to-action feel intentional and are paired with persuasive proof.",
        "gap": "Forms and CTAs lack urgency, so visitors stall before taking action.",
        "action": "Design a focused conversion path with bold CTAs, risk reducers, and social proof in strategic locations.",
    },
    "Accessibility": {
        "strength": "Color contrast and interaction states support inclusive browsing.",
        "gap": "Contrast, keyboard focus, or alt text gaps will create friction for a growing portion of your audience.",
        "action": "Address contrast ratios, alt text, and focus states to align with WCAG AA expectations.",
    },
}


def create_empty_client_dataframe() -> pd.DataFrame:
    """Return an empty client dataframe with the expected schema."""

    return pd.DataFrame({col: pd.Series(dtype=dtype) for col, dtype in CLIENT_DATA_SCHEMA.items()})


def ensure_client_dataframe_schema(df: pd.DataFrame) -> pd.DataFrame:
    """Guarantee that *df* contains every column defined in the schema."""

    for column, dtype in CLIENT_DATA_SCHEMA.items():
        if column not in df.columns:
            df[column] = pd.Series(dtype=dtype)
    return df


def _human_join(items: Sequence[str]) -> str:
    """Return a human-friendly comma-separated list."""

    if not items:
        return ""
    if len(items) == 1:
        return items[0]
    return ", ".join(items[:-1]) + f" and {items[-1]}"


def _chunk_sequence(items: Sequence, size: int) -> Iterable[Sequence]:
    """Yield *items* in chunks of *size* elements."""

    for start in range(0, len(items), size):
        yield items[start : start + size]


def _normalize_collection(value) -> List[str]:
    """Convert stored session data into a list of bullet points."""

    if value is None or (isinstance(value, float) and np.isnan(value)):
        return []
    if isinstance(value, list):
        return [str(item) for item in value if str(item).strip()]
    if isinstance(value, str):
        parts = [part.strip() for part in value.split("|") if part.strip()]
        if parts:
            return parts
        # Fallback: attempt to split comma-separated strings
        return [segment.strip() for segment in value.split(",") if segment.strip()]
    return [str(value)]


def _parse_breakdown(value) -> Dict[str, float]:
    """Best-effort parsing for stored design breakdown data."""

    if isinstance(value, dict):
        return {str(key): float(val) for key, val in value.items()}
    if isinstance(value, str) and value.strip():
        try:
            parsed = ast.literal_eval(value)
        except (ValueError, SyntaxError):
            return {}
        if isinstance(parsed, dict):
            try:
                return {str(key): float(val) for key, val in parsed.items()}
            except (TypeError, ValueError):
                return {}
    return {}


def _prepare_filter_options(series: pd.Series) -> List[str]:
    """Return sorted filter options without null values."""

    unique_values = series.unique()
    cleaned_values = [value for value in unique_values if pd.notna(value)]
    return sorted(str(value) for value in cleaned_values)


# Load custom CSS from external file
def load_css(css_file: str) -> str:
    css_path = get_asset_path(css_file)
    try:
        return css_path.read_text(encoding="utf-8")
    except FileNotFoundError:
        st.warning("Custom styles could not be loaded; falling back to defaults.")
        return ""

# Apply the custom CSS
st.markdown(f"<style>{load_css('styles.css')}</style>", unsafe_allow_html=True)

# App title and description
st.title("Web Redesign Client Scouting Tool")
st.subheader("Track and analyze potential clients for your web redesign business")

# Sidebar for navigation with improved logo
st.sidebar.markdown(
    """
    <div class="sidebar-brand">
        <div class="sidebar-brand__mark">Scout</div>
        <div class="sidebar-brand__subtitle">Client Intelligence</div>
    </div>
    """,
    unsafe_allow_html=True,
)
st.sidebar.markdown("<div style='margin-bottom:1.5rem;'></div>", unsafe_allow_html=True)
page = st.sidebar.selectbox("Navigation", ["Dashboard", "Client Database", "Website Analyzer", "Export Data", "Settings"])

# Function to calculate website age in years
def calculate_age(date):
    if pd.isna(date):
        return np.nan
    return (datetime.now() - pd.to_datetime(date)).days / 365.25

# Function to analyze a website using live heuristics
def analyze_website(url):
    def _normalize_url(raw_url: str) -> str:
        candidate = raw_url.strip()
        if not candidate:
            raise ValueError("Please provide a website URL to analyze.")
        parsed = urlparse(candidate)
        if not parsed.scheme:
            candidate = f"https://{candidate}"
            parsed = urlparse(candidate)
        if not parsed.netloc:
            raise ValueError("The website URL is missing a domain name.")
        return candidate

    def _fetch(session: requests.Session, target_url: str):
        start = time.perf_counter()
        response = session.get(target_url, timeout=12)
        elapsed_ms = (time.perf_counter() - start) * 1000
        return response, elapsed_ms

    def _clamp(value: float, lower: float = 0, upper: float = 100) -> float:
        return max(lower, min(upper, value))

    def _score_response_time(ms: float) -> float:
        if ms <= 1000:
            return 95
        if ms <= 2000:
            return 85
        if ms <= 3000:
            return 70
        if ms <= 4500:
            return 55
        if ms <= 6000:
            return 45
        return 35

    def _score_page_weight(kb: float) -> float:
        if kb <= 700:
            return 92
        if kb <= 1500:
            return 80
        if kb <= 2500:
            return 65
        if kb <= 4000:
            return 52
        return 40

    def _score_resource_count(count: int) -> float:
        if count <= 25:
            return 90
        if count <= 40:
            return 75
        if count <= 60:
            return 60
        return 45

    try:
        normalized_url = _normalize_url(url)
    except ValueError as exc:
        st.error(str(exc))
        return None

    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/122.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
        }
    )

    try:
        response, response_time_ms = _fetch(session, normalized_url)
    except requests.RequestException:
        if normalized_url.startswith("https://"):
            fallback = "http://" + normalized_url[len("https://") :]
            try:
                response, response_time_ms = _fetch(session, fallback)
                normalized_url = fallback
            except requests.RequestException as exc:  # pragma: no cover - network errors are surfaced to the UI
                st.error(f"Failed to reach {url}. {exc}")
                return None
        else:
            st.error(f"Failed to reach {url}. Please verify the address and try again.")
            return None
    except Exception as exc:  # pragma: no cover - safety net for unexpected issues
        st.error(f"Failed to analyze {url}. {exc}")
        return None

    status_code = response.status_code
    content = response.content or b""
    page_size_kb = len(content) / 1024 if content else 0.0

    if not response.text:
        response.encoding = response.apparent_encoding or "utf-8"
    page_text = response.text or content.decode("utf-8", errors="ignore")

    soup = BeautifulSoup(page_text, "html.parser")

    last_modified_header = response.headers.get("Last-Modified")
    last_update: datetime | None
    if last_modified_header:
        parsed_last_update = pd.to_datetime(last_modified_header, utc=True, errors="coerce")
        if pd.notna(parsed_last_update):
            last_update = parsed_last_update.tz_convert(None).to_pydatetime()
        else:
            last_update = None
    else:
        last_update = None

    meta_viewport = soup.find("meta", attrs={"name": re.compile("viewport", re.I)})
    if not meta_viewport:
        meta_viewport = soup.find("meta", attrs={"property": re.compile("viewport", re.I)})
    viewport_content = meta_viewport.get("content", "").lower() if meta_viewport else ""
    mobile_friendly = "width=device-width" in viewport_content or "initial-scale" in viewport_content

    images = soup.find_all("img")
    image_count = len(images)
    images_with_alt = sum(1 for img in images if (img.get("alt") or "").strip())
    missing_alt_count = image_count - images_with_alt
    alt_ratio = images_with_alt / image_count if image_count else 1.0

    scripts = [script for script in soup.find_all("script") if script.get("src")]
    script_count = len(scripts)

    links = soup.find_all("a")
    cta_keywords = [
        "contact",
        "book",
        "schedule",
        "demo",
        "quote",
        "start",
        "consult",
        "call",
        "enquire",
        "enquiry",
        "enroll",
        "buy",
        "shop",
        "signup",
        "sign up",
        "get started",
    ]
    cta_count = 0
    for link in links:
        text = link.get_text(strip=True).lower()
        if text and any(keyword in text for keyword in cta_keywords):
            cta_count += 1
    tel_links = sum(
        1 for link in links if (link.get("href") or "").lower().startswith(("tel:", "mailto:"))
    )

    forms_count = len(soup.find_all("form"))

    text_content = " ".join(segment.strip() for segment in soup.stripped_strings)
    word_count = len(text_content.split())
    paragraphs = soup.find_all("p")
    paragraph_count = len(paragraphs)
    avg_paragraph_words = word_count / paragraph_count if paragraph_count else float(word_count)

    structured_data = soup.find_all(
        "script", attrs={"type": lambda value: value and "ld+json" in value.lower()}
    )

    response_score = _score_response_time(response_time_ms)
    weight_score = _score_page_weight(page_size_kb)
    resource_score = _score_resource_count(image_count + script_count)
    speed_score = int(round(_clamp(0.5 * response_score + 0.3 * weight_score + 0.2 * resource_score)))

    title_tag = soup.find("title")
    title_length = len(title_tag.get_text(strip=True)) if title_tag else 0
    favicon_present = bool(
        soup.find("link", rel=lambda value: value and "icon" in value.lower())
    )
    og_site_name = bool(soup.find("meta", attrs={"property": "og:site_name"}))

    brand_score = 60
    if favicon_present:
        brand_score += 8
    if og_site_name:
        brand_score += 7
    if title_length >= 30:
        brand_score += 5
    if not title_tag:
        brand_score -= 12
    brand_score = _clamp(brand_score, 35, 92)

    visual_score = 62
    heading_counts = {tag: len(soup.find_all(tag)) for tag in ["h1", "h2", "h3"]}
    if heading_counts.get("h1", 0) == 1:
        visual_score += 10
    elif heading_counts.get("h1", 0) == 0:
        visual_score -= 12
    if heading_counts.get("h2", 0) >= 2:
        visual_score += 8
    if heading_counts.get("h3", 0) >= 3:
        visual_score += 4
    if paragraph_count and avg_paragraph_words <= 110:
        visual_score += 6
    elif avg_paragraph_words > 150:
        visual_score -= 8
    visual_score = _clamp(visual_score, 30, 90)

    content_score = 65
    if 400 <= word_count <= 1500:
        content_score += 6
    elif word_count > 2200 or word_count < 150:
        content_score -= 8
    if paragraph_count >= 8:
        content_score += 4
    if avg_paragraph_words < 80:
        content_score += 4
    elif avg_paragraph_words > 140:
        content_score -= 5
    if structured_data:
        content_score += 4
    content_score = _clamp(content_score, 35, 92)

    conversion_score = 55
    if cta_count >= 3:
        conversion_score += 15
    elif cta_count >= 1:
        conversion_score += 8
    else:
        conversion_score -= 6
    if forms_count >= 1:
        conversion_score += 10
    if tel_links >= 1:
        conversion_score += 5
    conversion_score = _clamp(conversion_score, 30, 90)

    accessibility_score = 68
    if alt_ratio >= 0.8:
        accessibility_score += 8
    elif alt_ratio < 0.5:
        accessibility_score -= 10
    if mobile_friendly:
        accessibility_score += 6
    else:
        accessibility_score -= 12
    if heading_counts.get("h1", 0) == 1:
        accessibility_score += 4
    elif heading_counts.get("h1", 0) == 0:
        accessibility_score -= 6
    html_tag = soup.find("html")
    if html_tag and html_tag.get("lang"):
        accessibility_score += 4
    accessibility_score = _clamp(accessibility_score, 30, 92)

    design_breakdown = {
        "Brand Cohesion": round(brand_score),
        "Visual Hierarchy": round(visual_score),
        "Content Clarity": round(content_score),
        "Conversion Readiness": round(conversion_score),
        "Accessibility": round(accessibility_score),
    }
    design_score = int(np.clip(np.mean(list(design_breakdown.values())), 0, 100))

    strengths: List[str] = []
    gaps: List[str] = []
    recommended_actions: List[str] = []

    for category, score in design_breakdown.items():
        details = DESIGN_CATEGORY_LIBRARY.get(category, {})
        if score >= 72:
            strengths.append(f"{category} ({score}/100) — {details.get('strength', '')}")
        elif score <= 65:
            gaps.append(f"{category} ({score}/100) — {details.get('gap', '')}")
            recommended_actions.append(
                f"Raise {category.lower()} by {details.get('action', 'designing focused improvements for this area.')}"
            )
        else:
            recommended_actions.append(
                f"Tighten {category.lower()} ({score}/100) so it matches the stronger sections. {details.get('action', '')}"
            )

    recommended_actions = list(dict.fromkeys([action for action in recommended_actions if action.strip()]))
    strengths = list(dict.fromkeys([item for item in strengths if item.strip()]))
    gaps = list(dict.fromkeys([item for item in gaps if item.strip()]))

    weakest = [name for name, value in sorted(design_breakdown.items(), key=lambda kv: kv[1])[:2]]
    strongest = [name for name, value in sorted(design_breakdown.items(), key=lambda kv: kv[1], reverse=True)[:2]]

    site_age = calculate_age(last_update)

    summary_parts = [
        f"The design benchmark for {normalized_url} lands at **{design_score}/100** based on live structure, content, and accessibility checks."
    ]
    if pd.notna(site_age):
        summary_parts.append(
            f"Server headers suggest the last significant refresh was roughly {site_age:.1f} years ago, which shapes how modern the experience feels."
        )
    summary_parts.append(
        f"The page responded in {response_time_ms/1000:.1f}s and weighs {page_size_kb:.0f} KB, two signals prospects feel immediately."
    )
    if weakest:
        summary_parts.append(
            f"Greatest friction sits within {_human_join(weakest)} where cohesion and storytelling taper off."
        )
    if strongest:
        summary_parts.append(
            f"Strengths you can amplify include {_human_join(strongest)}."
        )
    summary_parts.append(
        "Use this blend of hard metrics and narrative talking points to frame the redesign opportunity with clients."
    )
    design_summary = " ".join(summary_parts)

    if not gaps:
        gaps.append(
            "The core system is strong; focus on polishing micro-interactions to stay ahead of competitors."
        )
    if not recommended_actions:
        recommended_actions.append(
            "Document a lightweight design system playbook to protect the gains made across the experience."
        )

    evidence_points: List[str] = []
    evidence_points.append(
        f"First response landed in {response_time_ms:.0f} ms with a {status_code} status code — buyers expect <1500 ms for a premium feel."
    )
    evidence_points.append(
        f"Page weight is {page_size_kb:.0f} KB across {image_count} images and {script_count} scripts, which influences load speed and perception."
    )
    if image_count and missing_alt_count:
        evidence_points.append(
            f"{missing_alt_count} of {image_count} images are missing alt text, leaving accessibility and SEO equity on the table."
        )
    if cta_count == 0:
        evidence_points.append("No primary calls-to-action were detected, so visitors lack a clear next step.")
    elif cta_count < 2:
        evidence_points.append(
            f"Only {cta_count} clear call-to-action link{'s' if cta_count != 1 else ''} were detected, limiting conversion paths."
        )
    if forms_count == 0:
        evidence_points.append("No lead capture forms or booking widgets were present on the scanned page.")
    if pd.notna(site_age):
        evidence_points.append(
            f"Server headers indicate a refresh cadence of about {site_age:.1f} years, signalling dated conventions."
        )
    evidence_points = list(dict.fromkeys([point for point in evidence_points if point.strip()]))

    client_value_points: List[str] = []
    if response_time_ms > 1800:
        client_value_points.append(
            f"Improving load time from {response_time_ms/1000:.1f}s toward 1.5s can recover 5–10% of stalled conversions."
        )
    if page_size_kb > 1500:
        client_value_points.append(
            "Optimising imagery and code to lighten the page will reduce bounce rates on mobile traffic."
        )
    if not mobile_friendly:
        client_value_points.append(
            "Mobile visitors encounter friction — 65% of B2B research now happens on phones, so this is high-impact."
        )
    if conversion_score < 70 or cta_count < 2:
        client_value_points.append(
            "Designing a focused conversion path with more prominent CTAs typically lifts booked calls by double digits."
        )
    if accessibility_score < 70 or alt_ratio < 0.8:
        client_value_points.append(
            "Closing accessibility gaps protects compliance and unlocks enterprise procurement opportunities."
        )
    if not client_value_points:
        client_value_points.append(
            "Even with respectable fundamentals, a strategic refresh can showcase new offerings and keep the brand ahead of competitors."
        )

    site_age_years = site_age if pd.notna(site_age) else np.nan

    alt_coverage_value = (
        f"{images_with_alt}/{image_count} images with alt text" if image_count else "No images detected"
    )

    technical_metrics = [
        {
            "label": "Response Time",
            "value": f"{response_time_ms:.0f} ms",
            "client_context": "Aim for under 1500 ms so prospects feel the site respond instantly.",
        },
        {
            "label": "Page Weight",
            "value": f"{page_size_kb:.0f} KB",
            "client_context": "Keeping pages under ~1500 KB helps mobile visitors stay engaged.",
        },
        {
            "label": "Status Code",
            "value": str(status_code),
            "client_context": "Reliable 200 responses build trust that the experience is stable.",
        },
        {
            "label": "Image Alt Coverage",
            "value": alt_coverage_value,
            "client_context": "Alt text supports accessibility compliance and search visibility.",
        },
        {
            "label": "Clear CTAs",
            "value": str(cta_count),
            "client_context": "We target 2–3 high-intent CTAs per key page to drive action.",
        },
        {
            "label": "Lead Capture Points",
            "value": f"{forms_count} forms",
            "client_context": "Forms or booking widgets convert interest into conversations.",
        },
    ]

    return {
        'mobile_friendly': mobile_friendly,
        'speed_score': speed_score,
        'design_score': design_score,
        'last_update': last_update if last_update else pd.NaT,
        'design_breakdown': design_breakdown,
        'design_summary': design_summary,
        'design_strengths': strengths,
        'design_gaps': gaps,
        'recommended_actions': recommended_actions,
        'client_value_points': client_value_points,
        'evidence_points': evidence_points,
        'response_time_ms': response_time_ms,
        'page_size_kb': page_size_kb,
        'status_code': status_code,
        'image_count': image_count,
        'script_count': script_count,
        'word_count': word_count,
        'cta_count': cta_count,
        'forms_count': forms_count,
        'technical_metrics': technical_metrics,
        'site_age_years': site_age_years,
        'alt_ratio': alt_ratio,
        'missing_alt': missing_alt_count,
    }
    

# Function to export to Excel with formatting
def export_to_excel(df, filename="client_scouting_data.xlsx"):
    # Create a path for the file in the current directory
    filepath = Path(filename)
    
    # Convert DataFrame to Excel
    df.to_excel(filepath, index=False)
    
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Define styles
    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styles
    for col in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # Adjust column width based on content
        column_letter = get_column_letter(col)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Apply alternating row colors and borders to data cells
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            # Apply light gray fill to alternate rows
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Add a chart for website speed scores
    chart = BarChart()
    chart.title = "Website Speed Scores"
    chart.y_axis.title = "Company"
    chart.x_axis.title = "Speed Score"
    
    # Find the columns for company name and speed score
    company_col = df.columns.get_loc("Company Name") + 1
    speed_col = df.columns.get_loc("Website Speed Score") + 1
    
    # Create chart data references
    data = Reference(ws, min_col=speed_col, min_row=2, max_row=len(df) + 1)
    cats = Reference(ws, min_col=company_col, min_row=2, max_row=len(df) + 1)
    
    chart.add_data(data)
    chart.set_categories(cats)
    
    # Add the chart to the worksheet
    ws.add_chart(chart, f"A{len(df) + 5}")
    
    # Save the workbook
    wb.save(filepath)
    
    return filepath

# Initialize session state for client data if it doesn't exist
if 'client_data' not in st.session_state:
    st.session_state.client_data = create_empty_client_dataframe()

st.session_state.client_data = ensure_client_dataframe_schema(st.session_state.client_data)

# Ensure datetime columns remain consistent across reruns
for column in ["Last Website Update", "Last Contact Date"]:
    st.session_state.client_data[column] = pd.to_datetime(
        st.session_state.client_data[column], errors='coerce'
    )

# Dashboard Page
if page == "Dashboard":
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.subheader("Client Overview")
        
        # Calculate current metrics
        total_clients = len(st.session_state.client_data)
        avg_website_age = (
            st.session_state.client_data['Last Website Update']
            .apply(calculate_age)
            .dropna()
            .mean()
        )
        total_potential_value = st.session_state.client_data['Potential Value'].sum()
        mobile_unfriendly_count = st.session_state.client_data['Mobile Friendly'].value_counts().get(False, 0)
        
        # Display metrics in a grid with improved styling
        with styled_card():
            metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)

            with metric_col1:
                st.metric("Total Prospects", f"{total_clients}")

            with metric_col2:
                avg_age_display = f"{avg_website_age:.1f} years" if total_clients else "—"
                st.metric("Avg Website Age", avg_age_display)

            with metric_col3:
                st.metric(
                    "Total Potential Value",
                    f"${total_potential_value:,.0f}" if total_clients else "—",
                )

            with metric_col4:
                st.metric(
                    "Not Mobile Friendly",
                    f"{mobile_unfriendly_count} sites" if total_clients else "—",
                )

        # Create charts with improved styling
        st.subheader("Analysis")
        perf_tab, value_tab = st.tabs(["Performance Snapshot", "Value vs. Website Age"])

        with perf_tab:
            if total_clients:
                with styled_card():
                    fig1 = px.bar(
                        st.session_state.client_data,
                        x='Company Name',
                        y='Website Speed Score',
                        color='Priority',
                        color_discrete_map={'High': '#e53e3e', 'Medium': '#ed8936', 'Low': '#38a169'},
                        title="Website Speed Score by Company"
                    )
                    fig1.update_layout(
                        height=400,
                        plot_bgcolor='rgba(0,0,0,0)',
                        paper_bgcolor='rgba(0,0,0,0)',
                        title_font=dict(size=18, color='#e2e8f0'),
                        font=dict(family="Inter, Segoe UI, sans-serif", color='#94a3b8'),
                        margin=dict(l=40, r=40, t=60, b=40),
                    )
                    st.plotly_chart(fig1, use_container_width=True)
            else:
                with styled_card():
                    st.info("Add prospects to visualize performance trends.")

        with value_tab:
            if total_clients:
                with styled_card():
                    fig2 = px.scatter(
                        st.session_state.client_data,
                        x=st.session_state.client_data['Last Website Update'].apply(calculate_age),
                        y='Potential Value',
                        size='Design Score',
                        color='Industry',
                        hover_name='Company Name',
                        title="Potential Value vs. Website Age"
                    )
                    fig2.update_xaxes(title="Website Age (Years)")
                    fig2.update_yaxes(title="Potential Value ($)")
                    fig2.update_layout(
                        height=400,
                        plot_bgcolor='rgba(0,0,0,0)',
                        paper_bgcolor='rgba(0,0,0,0)',
                        title_font=dict(size=18, color='#e2e8f0'),
                        font=dict(family="Inter, Segoe UI, sans-serif", color='#94a3b8'),
                        margin=dict(l=40, r=40, t=60, b=40),
                    )
                    st.plotly_chart(fig2, use_container_width=True)
            else:
                with styled_card():
                    st.info("Once you add clients you'll see value potential plotted here.")
    
    with col2:
        with styled_card():
            st.subheader("Priority Prospects")

            # Filter high priority prospects
            high_priority = st.session_state.client_data[st.session_state.client_data['Priority'] == 'High'].sort_values('Potential Value', ascending=False)

            if len(high_priority) > 0:
                for _, client in high_priority.iterrows():
                    with st.container():
                        st.markdown(f"""
                        <div class="client-card high-priority">
                            <h4>{client['Company Name']}</h4>
                            <p style='color:#718096; font-size:0.9rem;'>{client['Industry']}</p>
                            <p><span style='font-weight:500;'>Contact:</span> {client['Contact Person']}</p>
                            <p><span style='font-weight:500;'>Value:</span> ${client['Potential Value']:,.0f}</p>
                            <p><span style='font-weight:500;'>Status:</span> <span class="status-badge status-high">{client['Status']}</span></p>
                        </div>
                        """, unsafe_allow_html=True)
            else:
                st.info("No high priority prospects yet. Add some in the Client Database.")

        with styled_card():
            st.subheader("Recent Activity")

            # Sort by most recent contact
            recent_contacts = st.session_state.client_data.sort_values('Last Contact Date', ascending=False).head(5)

            if len(recent_contacts) > 0:
                for _, client in recent_contacts.iterrows():
                    days_ago = (datetime.now() - pd.to_datetime(client['Last Contact Date'])).days
                    priority_class = "high-priority" if client['Priority'] == "High" else "medium-priority" if client['Priority'] == "Medium" else "low-priority"
                    status_badge_class = "status-high" if client['Priority'] == "High" else "status-medium" if client['Priority'] == "Medium" else "status-low"

                    st.markdown(f"""
                    <div class="client-card {priority_class}" style="padding:0.75rem; margin-bottom:0.75rem;">
                        <div style="display:flex; justify-content:space-between; align-items:center;">
                            <div>
                                <p style="margin:0; font-weight:600;">{client['Company Name']}</p>
                                <p style="margin:0;">
                                    <span class="status-badge {status_badge_class}">{client['Status']}</span>
                                </p>
                            </div>
                            <div class="activity-badge">
                                {days_ago} days ago
                            </div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.info("No recent activity yet.")

# Client Database Page
elif page == "Client Database":
    st.subheader("Client Database")
    
    # Filters in expandable section
    with st.expander("Filters"):
        col1, col2, col3 = st.columns(3)

        with col1:
            industry_options = _prepare_filter_options(st.session_state.client_data['Industry'])
            industry_filter = st.multiselect(
                "Industry",
                options=industry_options,
                default=[]
            )

        with col2:
            priority_options = _prepare_filter_options(st.session_state.client_data['Priority'])
            priority_filter = st.multiselect(
                "Priority",
                options=priority_options,
                default=[]
            )

        with col3:
            status_options = _prepare_filter_options(st.session_state.client_data['Status'])
            status_filter = st.multiselect(
                "Status",
                options=status_options,
                default=[]
            )

    # Apply filters
    filtered_data = st.session_state.client_data.copy()

    if industry_filter:
        filtered_data = filtered_data[filtered_data['Industry'].astype(str).isin(industry_filter)]

    if priority_filter:
        filtered_data = filtered_data[filtered_data['Priority'].astype(str).isin(priority_filter)]

    if status_filter:
        filtered_data = filtered_data[filtered_data['Status'].astype(str).isin(status_filter)]
    
    # Add new client form
    with st.expander("Add New Client"):
        with st.form("new_client_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                new_company = st.text_input("Company Name")
                new_website = st.text_input("Website URL")
                new_industry = st.selectbox("Industry", options=[
                    "Technology", "Manufacturing", "Software", "Retail", "Services", 
                    "Healthcare", "Education", "Finance", "Entertainment", "Other"
                ])
                new_contact = st.text_input("Contact Person")
            
            with col2:
                new_email = st.text_input("Contact Email")
                new_phone = st.text_input("Contact Phone")
                new_priority = st.selectbox("Priority", options=["High", "Medium", "Low"])
                new_status = st.selectbox("Status", options=[
                    "Prospecting", "Initial Contact", "Meeting Scheduled", 
                    "Proposal Sent", "Negotiation", "Closed Won", "Closed Lost"
                ])
            
            new_notes = st.text_area("Notes")
            
            submitted = st.form_submit_button("Add Client")
            
            if submitted and new_company and new_website:
                # Analyze the website
                st.info("Analyzing website... This may take a moment.")
                analysis = analyze_website(new_website)
                
                if analysis:
                    # Create new client record
                    new_client = {
                        'Company Name': new_company,
                        'Website URL': new_website,
                        'Industry': new_industry,
                        'Contact Person': new_contact,
                        'Contact Email': new_email,
                        'Contact Phone': new_phone,
                        'Last Website Update': analysis['last_update'],
                        'Mobile Friendly': analysis['mobile_friendly'],
                        'Website Speed Score': analysis['speed_score'],
                        'Design Score': analysis['design_score'],
                        'Design Summary': analysis['design_summary'],
                        'Design Strengths': analysis['design_strengths'],
                        'Design Gaps': analysis['design_gaps'],
                        'Design Recommendations': analysis['recommended_actions'],
                        'Design Breakdown': analysis['design_breakdown'],
                        'Potential Value': random.randint(25000, 200000),  # Random value for demo
                        'Priority': new_priority,
                        'Notes': new_notes,
                        'Last Contact Date': datetime.now(),
                        'Status': new_status
                    }
                    
                    # Add to session state
                    st.session_state.client_data = pd.concat([
                        st.session_state.client_data, 
                        pd.DataFrame([new_client])
                    ], ignore_index=True)
                    
                    st.success(f"Added {new_company} to the client database!")
                    st.experimental_rerun()
    
    if filtered_data.empty:
        with styled_card():
            st.info("Add prospects from the analyzer to build your working list.")
    else:
        # Display the data table with edit capability
        with styled_card():
            st.dataframe(
                filtered_data,
                use_container_width=True,
                height=400
            )

        # Action buttons for selected client
        st.subheader("Client Actions")
        selected_client_data = None
        with styled_card():
            selected_client = st.selectbox("Select Client", options=filtered_data['Company Name'].tolist())

            if selected_client:
                selected_client_data = filtered_data[filtered_data['Company Name'] == selected_client].iloc[0]

                col1, col2, col3, col4 = st.columns(4)

                with col1:
                    if st.button("Edit Client"):
                        st.session_state.edit_client = selected_client

                with col2:
                    if st.button("Delete Client"):
                        st.session_state.client_data = st.session_state.client_data[
                            st.session_state.client_data['Company Name'] != selected_client
                        ]
                        st.success(f"Deleted {selected_client} from the database.")
                        st.experimental_rerun()

                with col3:
                    if st.button("Log Contact"):
                        # Update last contact date
                        client_index = st.session_state.client_data.index[
                            st.session_state.client_data['Company Name'] == selected_client
                        ].tolist()[0]
                        st.session_state.client_data.at[client_index, 'Last Contact Date'] = datetime.now()
                        st.success(f"Updated last contact date for {selected_client}.")
                        st.experimental_rerun()

                with col4:
                    if st.button("Update Status"):
                        st.session_state.update_status_client = selected_client

            else:
                selected_client_data = None

        if 'update_status_client' in st.session_state:
            target_client = st.session_state.update_status_client
            if target_client in filtered_data['Company Name'].values:
                client_data = st.session_state.client_data[st.session_state.client_data['Company Name'] == target_client].iloc[0]
                with styled_card():
                    with st.form("update_status_form"):
                        new_status = st.selectbox(
                            "New Status",
                            options=[
                                "Prospecting", "Initial Contact", "Meeting Scheduled",
                                "Proposal Sent", "Negotiation", "Closed Won", "Closed Lost"
                            ],
                            index=[
                                "Prospecting", "Initial Contact", "Meeting Scheduled",
                                "Proposal Sent", "Negotiation", "Closed Won", "Closed Lost"
                            ].index(client_data['Status'])
                        )
                        notes = st.text_area("Status Update Notes")

                        if st.form_submit_button("Update"):
                            client_index = st.session_state.client_data.index[
                                st.session_state.client_data['Company Name'] == target_client
                            ].tolist()[0]

                            st.session_state.client_data.at[client_index, 'Status'] = new_status

                            if notes:
                                current_notes = st.session_state.client_data.at[client_index, 'Notes'] or ""
                                new_notes = f"{current_notes}\n\n[{datetime.now().strftime('%Y-%m-%d')}] Status changed to {new_status}: {notes}".strip()
                                st.session_state.client_data.at[client_index, 'Notes'] = new_notes

                            st.success(f"Updated status for {target_client} to {new_status}.")
                            del st.session_state.update_status_client
                            st.experimental_rerun()

        if 'edit_client' in st.session_state:
            target_client = st.session_state.edit_client
            if target_client in st.session_state.client_data['Company Name'].values:
                client_data = st.session_state.client_data[st.session_state.client_data['Company Name'] == target_client].iloc[0]
                with st.form("edit_client_form"):
                    st.subheader(f"Edit {target_client}")

                    col1, col2 = st.columns(2)

                    industries = [
                        "Technology", "Manufacturing", "Software", "Retail", "Services",
                        "Healthcare", "Education", "Finance", "Entertainment", "Other"
                    ]

                    with col1:
                        company = st.text_input("Company Name", value=client_data['Company Name'])
                        website = st.text_input("Website URL", value=client_data['Website URL'])
                        industry = st.selectbox(
                            "Industry",
                            options=industries,
                            index=industries.index(client_data['Industry']) if client_data['Industry'] in industries else 0,
                        )
                        contact = st.text_input("Contact Person", value=client_data['Contact Person'])

                    with col2:
                        email = st.text_input("Contact Email", value=client_data['Contact Email'])
                        phone = st.text_input("Contact Phone", value=client_data['Contact Phone'])
                        priority = st.selectbox(
                            "Priority",
                            options=["High", "Medium", "Low"],
                            index=["High", "Medium", "Low"].index(client_data['Priority'])
                        )
                        potential_value = st.number_input(
                            "Potential Value ($)",
                            min_value=0,
                            value=int(client_data['Potential Value']) if not pd.isna(client_data['Potential Value']) else 0,
                        )

                    notes = st.text_area("Notes", value=client_data['Notes'])

                    if st.form_submit_button("Save Changes"):
                        client_index = st.session_state.client_data.index[
                            st.session_state.client_data['Company Name'] == target_client
                        ].tolist()[0]

                        st.session_state.client_data.at[client_index, 'Company Name'] = company
                        st.session_state.client_data.at[client_index, 'Website URL'] = website
                        st.session_state.client_data.at[client_index, 'Industry'] = industry
                        st.session_state.client_data.at[client_index, 'Contact Person'] = contact
                        st.session_state.client_data.at[client_index, 'Contact Email'] = email
                        st.session_state.client_data.at[client_index, 'Contact Phone'] = phone
                        st.session_state.client_data.at[client_index, 'Priority'] = priority
                        st.session_state.client_data.at[client_index, 'Potential Value'] = potential_value
                        st.session_state.client_data.at[client_index, 'Notes'] = notes

                        st.success(f"Updated information for {company}.")
                        del st.session_state.edit_client
                        st.experimental_rerun()
        
        # Display client details with improved styling
        if selected_client_data is not None:
            with styled_card():
                st.subheader("Client Details")

                # Priority badge
                priority_color = "#e53e3e" if selected_client_data['Priority'] == "High" else "#ed8936" if selected_client_data['Priority'] == "Medium" else "#38a169"
                st.markdown(f"""
                <div style="margin-bottom:1rem;">
                    <span class="priority-badge" style="color:{priority_color}; border-color:{priority_color}33; background-color:{priority_color}1a;">
                        {selected_client_data['Priority']} Priority
                    </span>
                </div>
                """, unsafe_allow_html=True)

                col1, col2 = st.columns(2)

                with col1:
                    with st.container():
                        st.markdown("#### Contact Information")
                        col1a, col1b = st.columns([1, 2])
                        with col1a:
                            st.markdown("**Person:**")
                            st.markdown("**Email:**")
                            st.markdown("**Phone:**")
                        with col1b:
                            st.markdown(f"{selected_client_data['Contact Person']}")
                            st.markdown(f"{selected_client_data['Contact Email']}")
                            st.markdown(f"{selected_client_data['Contact Phone']}")

                        st.markdown("#### Business Information")
                        col1a, col1b = st.columns([1, 2])
                        with col1a:
                            st.markdown("**Industry:**")
                            st.markdown("**Potential Value:**")
                            st.markdown("**Status:**")
                            st.markdown("**Last Contact:**")
                        with col1b:
                            st.markdown(f"{selected_client_data['Industry']}")
                            st.markdown(f"${selected_client_data['Potential Value']:,.0f}")
                            st.markdown(f"{selected_client_data['Status']}")

                            last_contact_raw = selected_client_data.get('Last Contact Date')
                            if pd.isna(last_contact_raw):
                                last_contact_display = "—"
                            else:
                                last_contact_timestamp = pd.to_datetime(last_contact_raw)
                                last_contact_display = last_contact_timestamp.strftime('%Y-%m-%d')
                            st.markdown(last_contact_display)

                with col2:
                    with st.container():
                        st.markdown("#### Website Information")

                        col2a, col2b = st.columns([1, 2])
                        with col2a:
                            st.markdown("**URL:**")
                            st.markdown("**Last Update:**")
                            st.markdown("**Mobile Friendly:**")
                        with col2b:
                            st.markdown(f"[{selected_client_data['Website URL']}]({selected_client_data['Website URL']})")

                            last_update_raw = selected_client_data.get('Last Website Update')
                            if pd.isna(last_update_raw):
                                last_update_display = "—"
                            else:
                                last_update_timestamp = pd.to_datetime(last_update_raw)
                                age_years = calculate_age(last_update_timestamp)
                                if pd.isna(age_years):
                                    last_update_display = last_update_timestamp.strftime('%Y-%m-%d')
                                else:
                                    last_update_display = (
                                        f"{last_update_timestamp.strftime('%Y-%m-%d')} "
                                        f"({age_years:.1f} years ago)"
                                    )
                            st.markdown(last_update_display)
                            st.markdown(f"{'Yes' if selected_client_data['Mobile Friendly'] else 'No'}")

                        st.markdown("**Speed Score:**")
                        speed_score = selected_client_data['Website Speed Score']
                        st.progress(speed_score/100)
                        col2c1, col2c2 = st.columns([3, 1])
                        with col2c2:
                            st.markdown(f"{speed_score}/100")

                        st.markdown("**Design Score:**")
                        design_score = selected_client_data['Design Score']
                        st.progress(design_score/100)
                        col2d1, col2d2 = st.columns([3, 1])
                        with col2d2:
                            st.markdown(f"{design_score}/100")

                strengths = _normalize_collection(selected_client_data.get('Design Strengths'))
                gaps = _normalize_collection(selected_client_data.get('Design Gaps'))
                recommendations = _normalize_collection(selected_client_data.get('Design Recommendations'))

                design_summary_text = selected_client_data.get('Design Summary', '')
                if pd.notna(design_summary_text) and str(design_summary_text).strip():
                    st.markdown("#### Design Intelligence")
                    st.markdown(str(design_summary_text))

                    breakdown = _parse_breakdown(selected_client_data.get('Design Breakdown'))
                    if breakdown:
                        st.markdown("##### Score Breakdown")
                        for chunk in _chunk_sequence(list(breakdown.items()), 3):
                            metric_cols = st.columns(len(chunk))
                            for metric_col, (label, value) in zip(metric_cols, chunk):
                                with metric_col:
                                    st.metric(label, f"{value:.0f}/100")

                if strengths or gaps or recommendations:
                    insight_col1, insight_col2 = st.columns(2)
                    with insight_col1:
                        if strengths:
                            st.markdown("##### Strengths to Celebrate")
                            for item in strengths:
                                st.markdown(f"- ✅ {item}")
                        if gaps:
                            st.markdown("##### Friction Points")
                            for item in gaps:
                                st.markdown(f"- ⚠️ {item}")

                    with insight_col2:
                        if recommendations:
                            st.markdown("##### Strategic Next Steps")
                            for idx, action in enumerate(recommendations, 1):
                                st.markdown(f"{idx}. {action}")

            st.subheader("Notes")
            note_value = selected_client_data['Notes'] if pd.notna(selected_client_data['Notes']) else ""
            st.text_area(
                "Client Notes",
                value=str(note_value),
                height=200,
                key="readonly_notes",
                disabled=True,
            )

# Website Analyzer Page
elif page == "Website Analyzer":
    st.subheader("Website Analyzer")
    
    with st.form("analyzer_form"):
        website_url = st.text_input("Enter Website URL to Analyze")
        submit_button = st.form_submit_button("Analyze")
    
    if submit_button and website_url:
        with st.spinner("Analyzing website..."):
            analysis = analyze_website(website_url)
            
            if analysis:
                st.success("Analysis complete!")
                
                site_age_years = analysis.get('site_age_years')

                col1, col2, col3 = st.columns(3)

                with col1:
                    if site_age_years is not None and not pd.isna(site_age_years):
                        st.metric("Website Age", f"{site_age_years:.1f} years")
                    else:
                        st.metric("Website Age", "Unknown")

                with col2:
                    st.metric("Mobile Friendly", "Yes" if analysis['mobile_friendly'] else "No")

                with col3:
                    st.metric("Speed Score", f"{analysis['speed_score']}/100")

                technical_metrics = analysis.get('technical_metrics', [])
                if technical_metrics:
                    with styled_card("insight-card"):
                        st.markdown("#### Technical Performance Signals")
                        for chunk in _chunk_sequence(technical_metrics, 3):
                            metric_row = st.columns(len(chunk))
                            for metric_col, metric in zip(metric_row, chunk):
                                with metric_col:
                                    st.metric(metric['label'], metric['value'])
                                    st.caption(metric['client_context'])
                
                # Gauge chart for design score
                fig = go.Figure(go.Indicator(
                    mode = "gauge+number",
                    value = analysis['design_score'],
                    domain = {'x': [0, 1], 'y': [0, 1]},
                    title = {'text': "Design Score"},
                    gauge = {
                        'axis': {'range': [None, 100]},
                        'bar': {'color': "#38bdf8"},
                        'steps': [
                            {'range': [0, 40], 'color': "#7f1d1d"},
                            {'range': [40, 70], 'color': "#9a3412"},
                            {'range': [70, 100], 'color': "#065f46"}
                        ]
                    }
                ))
                
                st.plotly_chart(fig, use_container_width=True)

                st.markdown("### Design Intelligence Snapshot")

                breakdown_items = list(analysis['design_breakdown'].items())
                if breakdown_items:
                    with styled_card("insight-card"):
                        st.markdown("#### Score Breakdown")
                        for chunk in _chunk_sequence(breakdown_items, 3):
                            row_cols = st.columns(len(chunk))
                            for col, (label, value) in zip(row_cols, chunk):
                                with col:
                                    st.metric(label, f"{value:.0f}/100")
                                    st.progress(value / 100)

                story_col, value_col = st.columns([1.6, 1])

                with story_col:
                    with styled_card("insight-card"):
                        st.markdown("#### Storyline")
                        st.markdown(analysis['design_summary'])

                        strengths = analysis['design_strengths']
                        if strengths:
                            st.markdown("##### Strengths to Leverage")
                            for item in strengths:
                                st.markdown(f"- ✅ {item}")

                        gaps = analysis['design_gaps']
                        if gaps:
                            st.markdown("##### Where Users Feel Friction")
                            for item in gaps:
                                st.markdown(f"- ⚠️ {item}")

                with value_col:
                    with styled_card("insight-card accent-card"):
                        st.markdown("#### Client Value Talking Points")
                        for point in analysis['client_value_points']:
                            st.markdown(f"- 💡 {point}")

                    with styled_card("insight-card"):
                        st.markdown("#### Evidence to Share")
                        for point in analysis['evidence_points']:
                            st.markdown(f"- 📌 {point}")

                with styled_card("insight-card"):
                    st.markdown("#### High-Impact Next Steps")
                    for idx, action in enumerate(analysis['recommended_actions'], 1):
                        st.markdown(f"{idx}. {action}")

                # Opportunity assessment with improved styling
                st.subheader("Redesign Opportunity Assessment")

                opportunity_score = 0

                # Calculate opportunity score based on analysis
                if not analysis['mobile_friendly']:
                    opportunity_score += 30
                
                opportunity_score += (100 - analysis['speed_score']) * 0.3
                opportunity_score += (100 - analysis['design_score']) * 0.3
                site_age_for_score = (
                    site_age_years if site_age_years is not None and not pd.isna(site_age_years) else 0
                )
                opportunity_score += min(site_age_for_score * 5, 25)  # Cap at 25 points
                
                # Display recommendation based on score
                if opportunity_score > 70:
                    recommendation = "High Priority - This website is significantly outdated and presents an excellent opportunity for a complete redesign."
                    priority = "High"
                elif opportunity_score > 40:
                    recommendation = "Medium Priority - This website has several areas for improvement and would benefit from a redesign."
                    priority = "Medium"
                else:
                    recommendation = "Low Priority - This website is relatively modern but could still benefit from some improvements."
                    priority = "Low"
                
                # Use Streamlit native components for the opportunity assessment
                with styled_card("insight-card focus-card"):
                    # Header with priority badge
                    col1, col2 = st.columns([3, 1])
                    with col1:
                        st.subheader("Opportunity Assessment")
                    with col2:
                        st.markdown(f"**{priority} Priority**")
                    
                    # Opportunity score
                    st.markdown("**Opportunity Score**")
                    st.progress(opportunity_score/100)
                    
                    # Score value and scale
                    col_scale1, col_scale2, col_scale3 = st.columns(3)
                    with col_scale1:
                        st.markdown("0")
                    with col_scale2:
                        st.markdown("50")
                    with col_scale3:
                        st.markdown("100")
                    
                    # Score and recommendation
                    st.markdown(f"**{opportunity_score:.1f}/100**")
                    st.markdown(recommendation)
                    
                    # Key factors
                    st.markdown("**Key Factors:**")
                    factors = [
                        f"{'Not mobile-friendly' if not analysis['mobile_friendly'] else 'Mobile-friendly, but could be improved'}",
                        f"Speed score: {analysis['speed_score']}/100",
                        f"Design score: {analysis['design_score']}/100",
                        (
                            f"Website age: ~{site_age_years:.1f} years"
                            if site_age_years is not None and not pd.isna(site_age_years)
                            else "Website age: not available from headers"
                        ),
                    ]
                    breakdown = analysis.get('design_breakdown', {})
                    if breakdown:
                        weakest_focus = sorted(breakdown.items(), key=lambda kv: kv[1])[:2]
                        if weakest_focus:
                            focus_labels = [label for label, _ in weakest_focus]
                            factors.append(f"Design focus: {_human_join(focus_labels)} need attention to lift credibility.")
                    for factor in factors:
                        st.markdown(f"- {factor}")
                
                # Add to database option
                if st.button("Add to Client Database"):
                    # Check if already in database
                    if website_url in st.session_state.client_data['Website URL'].values:
                        st.warning("This website is already in your client database.")
                    else:
                        st.session_state.add_analyzed_site = {
                            'url': website_url,
                            'analysis': analysis
                        }
                        st.info("Please provide additional client details below.")
                
                # Form for adding the analyzed site
                if 'add_analyzed_site' in st.session_state:
                    with st.form("add_analyzed_client_form"):
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            company = st.text_input("Company Name")
                            industry = st.selectbox("Industry", options=[
                                "Technology", "Manufacturing", "Software", "Retail", "Services", 
                                "Healthcare", "Education", "Finance", "Entertainment", "Other"
                            ])
                            contact = st.text_input("Contact Person")
                        
                        with col2:
                            email = st.text_input("Contact Email")
                            phone = st.text_input("Contact Phone")
                            priority = st.selectbox("Priority", options=["High", "Medium", "Low"])
                        
                        notes = st.text_area("Notes")
                        
                        if st.form_submit_button("Add to Database"):
                            if company:
                                # Create new client record
                                new_client = {
                                    'Company Name': company,
                                    'Website URL': website_url,
                                    'Industry': industry,
                                    'Contact Person': contact,
                                    'Contact Email': email,
                                    'Contact Phone': phone,
                                    'Last Website Update': st.session_state.add_analyzed_site['analysis']['last_update'],
                                    'Mobile Friendly': st.session_state.add_analyzed_site['analysis']['mobile_friendly'],
                                    'Website Speed Score': st.session_state.add_analyzed_site['analysis']['speed_score'],
                                    'Design Score': st.session_state.add_analyzed_site['analysis']['design_score'],
                                    'Design Summary': st.session_state.add_analyzed_site['analysis']['design_summary'],
                                    'Design Strengths': st.session_state.add_analyzed_site['analysis']['design_strengths'],
                                    'Design Gaps': st.session_state.add_analyzed_site['analysis']['design_gaps'],
                                    'Design Recommendations': st.session_state.add_analyzed_site['analysis']['recommended_actions'],
                                    'Design Breakdown': st.session_state.add_analyzed_site['analysis']['design_breakdown'],
                                    'Potential Value': int(opportunity_score * 2000),  # Rough estimate based on opportunity score
                                    'Priority': priority,
                                    'Notes': notes,
                                    'Last Contact Date': datetime.now(),
                                    'Status': 'Prospecting'
                                }
                                
                                # Add to session state
                                st.session_state.client_data = pd.concat([
                                    st.session_state.client_data, 
                                    pd.DataFrame([new_client])
                                ], ignore_index=True)
                                
                                st.success(f"Added {company} to the client database!")
                                del st.session_state.add_analyzed_site
                                st.experimental_rerun()
                            else:
                                st.error("Company name is required.")

# Export Data Page
elif page == "Export Data":
    st.subheader("Export Client Data")
    
    # Preview the data to be exported
    st.dataframe(st.session_state.client_data, use_container_width=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Export options
        export_format = st.radio(
            "Export Format",
            options=["Excel (XLSX)", "CSV", "JSON"],
            index=0
        )
    
    with col2:
        # Filter options
        include_all = st.checkbox("Include all clients", value=True)
        
        if not include_all:
            export_priority = st.multiselect(
                "Include Priority Levels",
                options=["High", "Medium", "Low"],
                default=["High", "Medium", "Low"]
            )
        else:
            export_priority = ["High", "Medium", "Low"]
    
    # Filter data for export
    export_data = st.session_state.client_data[
        st.session_state.client_data['Priority'].isin(export_priority)
    ]
    
    # Add export button
    if st.button("Generate Export"):
        with st.spinner("Preparing export..."):
            if export_format == "Excel (XLSX)":
                # Export to formatted Excel
                filepath = export_to_excel(export_data)
                
                # Create download button
                with open(filepath, "rb") as file:
                    st.download_button(
                        label="Download Excel File",
                        data=file,
                        file_name="web_redesign_client_prospects.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            
            elif export_format == "CSV":
                # Export to CSV
                csv = export_data.to_csv(index=False)
                st.download_button(
                    label="Download CSV File",
                    data=csv,
                    file_name="web_redesign_client_prospects.csv",
                    mime="text/csv"
                )
            
            else:  # JSON
                # Export to JSON
                json_data = export_data.to_json(orient="records", date_format="iso")
                st.download_button(
                    label="Download JSON File",
                    data=json_data,
                    file_name="web_redesign_client_prospects.json",
                    mime="application/json"
                )
    
    # Export templates section
    st.subheader("Export Templates")
    st.write("Generate templated documents for your client outreach.")
    
    template_type = st.selectbox(
        "Select Template Type",
        options=["Initial Outreach Email", "Website Audit Report", "Proposal Document"]
    )

    client_options = export_data['Company Name'].tolist()
    if client_options:
        client_for_template = st.selectbox(
            "Select Client",
            options=client_options
        )

        if st.button("Generate Template"):
            client = export_data[export_data['Company Name'] == client_for_template].iloc[0]

            template_label = "Generated Template"
            download_label = "Download Template"
            download_name = f"{client_for_template.lower().replace(' ', '_')}_template.txt"
            download_mime = "text/plain"

            if template_type == "Initial Outreach Email":
                template_label = "Email Template"
                download_label = "Download Email Template"
                download_name = f"{client_for_template.lower().replace(' ', '_')}_outreach_email.txt"
                download_mime = "text/plain"
                template = f"""
Subject: Modernizing {client['Company Name']}'s Web Presence - Potential Partnership
            
Dear {client['Contact Person']},

I hope this email finds you well. My name is [Your Name] from [Your Company], and we specialize in redesigning and modernizing websites for businesses in the {client['Industry']} industry.

While reviewing industry websites, I noticed {client['Company Name']}'s site at {client['Website URL']} could benefit from some updates to align with current web standards and user expectations.

Some observations about your current website:
- It was last updated approximately {calculate_age(client['Last Website Update']):.1f} years ago
- {'It is not fully optimized for mobile devices' if not client['Mobile Friendly'] else 'While it has mobile support, the mobile experience could be improved'}
- The site's loading speed could be improved to enhance user experience and search engine rankings

I'd love to schedule a brief call to discuss how we could help modernize your web presence to better serve your business goals and attract more customers.

Would you be available for a 15-minute call next week to discuss potential improvements?

Best regards,
[Your Name]
[Your Company]
[Your Contact Information]
            """
            elif template_type == "Website Audit Report":
                template_label = "Audit Report Template"
                download_label = "Download Audit Report"
                download_name = f"{client_for_template.lower().replace(' ', '_')}_website_audit.md"
                download_mime = "text/markdown"
                template = f"""
# Website Audit Report for {client['Company Name']}
**Prepared by: [Your Company]**
**Date: {datetime.now().strftime('%B %d, %Y')}**

## Executive Summary
This audit examines the current state of {client['Company Name']}'s website ({client['Website URL']}) and identifies opportunities for improvement. The site was last significantly updated approximately {calculate_age(client['Last Website Update']):.1f} years ago, which suggests it may not incorporate current web design best practices and technologies.

## Technical Assessment

### Mobile Responsiveness: {'✓ Pass' if client['Mobile Friendly'] else '✗ Fail'}
{'The website is optimized for mobile devices.' if client['Mobile Friendly'] else 'The website is not fully optimized for mobile devices, which may negatively impact user experience for the growing number of mobile users and affect search rankings.'}

### Performance: {client['Website Speed Score']}/100
{'The website performs well and loads quickly.' if client['Website Speed Score'] > 80 else 'The website has performance issues that may lead to user frustration and abandonment.' if client['Website Speed Score'] > 50 else 'The website has significant performance issues that are likely causing poor user experience and affecting conversions.'}

### Design Assessment: {client['Design Score']}/100
{'The website has a modern, appealing design that aligns with current standards.' if client['Design Score'] > 80 else 'The website design appears dated and could benefit from modernization.' if client['Design Score'] > 50 else 'The website design is significantly outdated and does not meet current user expectations.'}

## Recommendations

Based on our assessment, we recommend the following improvements:

1. {'Maintain current mobile responsiveness, with minor tweaks to improve user experience.' if client['Mobile Friendly'] else 'Implement a fully responsive design that works seamlessly across all device types and screen sizes.'}

2. {'Optimize performance further to maintain competitive advantage.' if client['Website Speed Score'] > 80 else 'Address performance issues through code optimization, image compression, and modernized development practices.' if client['Website Speed Score'] > 50 else 'Complete overhaul of website architecture and codebase to address critical performance issues.'}

3. {'Refresh visual elements to maintain modern appearance.' if client['Design Score'] > 80 else 'Update visual design to align with current expectations and brand positioning.' if client['Design Score'] > 50 else 'Complete redesign to create a modern, engaging user experience that builds trust and drives conversions.'}

4. Update content to ensure it's fresh, relevant, and optimized for both users and search engines.

5. Implement current security standards to protect user data and maintain trust.

## Competitive Analysis
In your industry ({client['Industry']}), websites typically feature [industry-specific features]. Your competitors have implemented modern designs with [specific features], giving them a competitive advantage in user engagement and conversion rates.

## Next Steps
We would welcome the opportunity to discuss this audit in more detail and explore how we can help {client['Company Name']} implement these recommendations. Please contact us at [Your Contact Information] to schedule a consultation.
            """
            else:  # Proposal Document
                template_label = "Proposal Template"
                download_label = "Download Proposal Template"
                download_name = f"{client_for_template.lower().replace(' ', '_')}_website_proposal.md"
                download_mime = "text/markdown"
                template = f"""
# Website Redesign Proposal
**Prepared for: {client['Company Name']}**
**Contact: {client['Contact Person']}**
**Prepared by: [Your Company]**
**Date: {datetime.now().strftime('%B %d, %Y')}**

## Project Overview
[Your Company] is pleased to present this proposal to redesign the {client['Company Name']} website. Based on our analysis, your current website presents significant opportunities for improvement that can help drive business growth, improve user experience, and strengthen your online presence.

## Current Website Assessment
- **Last Major Update:** Approximately {calculate_age(client['Last Website Update']):.1f} years ago
- **Mobile Optimization:** {'Present but could be improved' if client['Mobile Friendly'] else 'Not optimized for mobile devices'}
- **Performance Score:** {client['Website Speed Score']}/100
- **Design Assessment:** {client['Design Score']}/100

## Proposed Solution
We propose a comprehensive website redesign that will address the identified issues and transform your online presence into a powerful business tool. Our solution includes:

1. **Modern, Responsive Design**
   - Fully responsive layout that works seamlessly across all devices
   - Custom design aligned with your brand identity
   - Intuitive navigation and user-friendly interface

2. **Performance Optimization**
   - Fast-loading pages optimized for both desktop and mobile
   - Efficient code structure and optimized assets
   - Implementation of best practices for web performance

3. **Content Strategy and SEO**
   - Content audit and restructuring for maximum impact
   - SEO optimization to improve search engine visibility
   - Compelling calls-to-action to drive user engagement

4. **Technology Stack**
   - Implementation of a modern, secure content management system
   - Integration with your existing business tools and systems
   - Scalable architecture to support future growth

## Project Timeline
- **Discovery Phase:** 2 weeks
- **Design Phase:** 3 weeks
- **Development Phase:** 4 weeks
- **Testing and Quality Assurance:** 1 week
- **Content Migration and Launch:** 2 weeks
- **Total Project Duration:** 12 weeks

## Investment
Based on the scope outlined above, the investment for this project is:

**Total Project Investment: $[Total Amount]**

Payment Schedule:
- 30% upon project initiation
- 30% upon design approval
- 40% upon project completion

## Why Choose [Your Company]
- Specialized experience in the {client['Industry']} industry
- Proven track record of successful website redesigns
- Dedicated project manager and support team
- Ongoing maintenance and support options
- Commitment to delivering measurable results

## Next Steps
To proceed with this project, please:
1. Review this proposal
2. Sign the attached agreement
3. Submit the initial payment
4. Schedule the kick-off meeting

We look forward to partnering with {client['Company Name']} to create a website that effectively represents your brand and drives business results.

[Your Signature]
[Your Name]
[Your Position]
[Your Company]
[Your Contact Information]
            """

            st.text_area(template_label, template, height=400)
            st.download_button(
                download_label,
                template,
                file_name=download_name,
                mime=download_mime
            )
    else:
        st.info("Add clients to the database to generate outreach templates.")

# Settings Page
elif page == "Settings":
    st.subheader("Application Settings")
    
    with st.expander("User Profile"):
        col1, col2 = st.columns(2)
        
        with col1:
            company_name = st.text_input("Your Company Name", value="Your Web Design Company")
            user_name = st.text_input("Your Name", value="John Doe")
        
        with col2:
            email = st.text_input("Your Email", value="contact@yourcompany.com")
            phone = st.text_input("Your Phone", value="555-123-4567")
        
        if st.button("Save Profile"):
            st.success("Profile information saved!")
    
    with st.expander("Industry Settings"):
        st.write("Customize the industries available in the application.")
        
        # Get current industries
        current_industries = sorted(st.session_state.client_data['Industry'].unique())
        
        # Allow adding new industries
        new_industry = st.text_input("Add New Industry")
        if st.button("Add Industry") and new_industry:
            if new_industry not in current_industries:
                st.success(f"Added {new_industry} to industries list.")
                current_industries.append(new_industry)
            else:
                st.warning(f"{new_industry} already exists in the industries list.")
        
        # Display and allow removal of industries
        st.write("Current Industries:")
        for industry in current_industries:
            col1, col2 = st.columns([3, 1])
            with col1:
                st.write(industry)
            with col2:
                if st.button(f"Remove {industry}", key=f"remove_{industry}"):
                    # Logic to remove would go here
                    st.success(f"Removed {industry} from industries list.")
    
    with st.expander("Data Management"):
        st.write("Manage your client data.")
        
        if st.button("Export All Data"):
            # Export all data to JSON
            json_data = st.session_state.client_data.to_json(orient="records", date_format="iso")
            st.download_button(
                label="Download All Data (JSON)",
                data=json_data,
                file_name="web_redesign_all_client_data.json",
                mime="application/json"
            )
        
        if st.button("Import Data"):
            st.info("Data import functionality would be implemented here.")
        
        if st.button("Clear All Data"):
            st.warning("Are you sure you want to clear all client data? This cannot be undone.")
            if st.button("Yes, Clear All Data", key="confirm_clear"):
                st.session_state.client_data = pd.DataFrame(columns=st.session_state.client_data.columns)
                st.success("All client data has been cleared.")
                st.experimental_rerun()
    
    with st.expander("Application Appearance"):
        st.write("Customize the appearance of the application.")
        
        theme_color = st.color_picker("Primary Theme Color", value="#3498db")
        
        accent_color = st.color_picker("Accent Color", value="#e74c3c")
        
        if st.button("Apply Theme"):
            st.success("Theme applied successfully!")
            # In a real app, this would update the CSS
            st.markdown(f"""
            <style>
            .stButton>button {{
                background-color: {theme_color};
            }}
            </style>
            """, unsafe_allow_html=True)

# Run the app
if __name__ == "__main__":
    # This script is designed to be run with Streamlit
    # Install required packages:
    # pip install streamlit pandas plotly openpyxl pillow requests beautifulsoup4
    # Run with: streamlit run web_redesign_client_scout.py
    pass
